"""
XLSX Parser - Parse Excel files with multiple sheets for post scheduling.
Each sheet = one scheduled run. Each row = one post for one profile.
Columns: title | url | community
"""
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def parse_xlsx(file_path: str) -> dict:
    """
    Parse XLSX file with support for multiple sheets.
    
    Structure:
    - Each sheet = one run/cycle
    - Each sheet has columns: title, url, community
    
    Returns:
    {
        "sheets": [
            {"sheet_name": "Sheet1", "index": 0, "posts": [...]},
            {"sheet_name": "Sheet2", "index": 1, "posts": [...]},
        ],
        "total_sheets": 2,
        "total_posts": 6
    }
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return {"sheets": [], "total_sheets": 0, "total_posts": 0, "error": "openpyxl not installed"}
    
    file_path = Path(file_path)
    
    if not file_path.is_file():
        logger.error(f"XLSX file not found: {file_path}")
        return {"sheets": [], "total_sheets": 0, "total_posts": 0, "error": "File not found"}
    
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open XLSX {file_path}: {e}")
        return {"sheets": [], "total_sheets": 0, "total_posts": 0, "error": str(e)}
    
    sheets = []
    total_posts = 0
    all_skipped = []
    
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), []):
            if cell:
                headers.append(str(cell).strip().lower())
        
        title_col = None
        url_col = None
        community_col = None
        
        for i, header in enumerate(headers):
            if 'title' in header:
                title_col = i
            elif 'url' in header or 'link' in header:
                url_col = i
            elif 'community' in header:
                community_col = i
        
        if title_col is None or url_col is None:
            logger.warning(f"Sheet '{sheet_name}': Missing title or url column, skipping")
            all_skipped.append({"sheet": sheet_name, "reason": "Missing title or url column"})
            continue
        
        posts = []
        skipped_rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                skipped_rows.append({"row": row_idx, "reason": "Empty row"})
                continue
            
            title = row[title_col] if len(row) > title_col else None
            url = row[url_col] if len(row) > url_col else None
            community = row[community_col] if community_col is not None and len(row) > community_col else None
            
            if not title or (isinstance(title, str) and not title.strip()):
                skipped_rows.append({"row": row_idx, "reason": "Missing or empty title"})
                continue
            
            url_str = str(url).strip() if url else ""
            if not url_str or not url_str.startswith(("http://", "https://")):
                skipped_rows.append({"row": row_idx, "reason": f"Invalid URL: '{url_str}'"})
                continue
            
            post_data = {
                "title": str(title).strip(),
                "url": url_str,
                "community": str(community).strip() if community and isinstance(community, str) and community.strip() else ""
            }
            posts.append(post_data)
        
        if skipped_rows:
            all_skipped.extend([{"sheet": sheet_name, **s} for s in skipped_rows])
        
        if posts:
            sheets.append({
                "sheet_name": sheet_name,
                "index": sheet_index,
                "posts": posts
            })
            total_posts += len(posts)
            logger.info(f"Sheet '{sheet_name}': {len(posts)} posts loaded, {len(skipped_rows)} skipped")
        else:
            logger.warning(f"Sheet '{sheet_name}': 0 valid posts (all rows skipped)")
    
    workbook.close()
    
    result = {
        "sheets": sheets,
        "total_sheets": len(sheets),
        "total_posts": total_posts,
        "skipped": all_skipped
    }
    
    logger.info(f"XLSX parsed: {len(sheets)} sheets, {total_posts} total posts, {len(all_skipped)} skipped")
    return result
